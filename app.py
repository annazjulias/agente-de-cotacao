import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# ==============================
# PADRONIZAÇÃO
# ==============================
def tratar_colunas(df):
    df.columns = (
        df.columns
        .str.strip()
        .str.lower()
        .str.replace(" ", "_")
        .str.replace("ç", "c")
        .str.replace("ã", "a")
        .str.replace("á", "a")
        .str.replace("é", "e")
        .str.replace("í", "i")
        .str.replace("ó", "o")
        .str.replace("ú", "u")
    )
    return df


def tratar_codigo(coluna):
    return (
        coluna.astype(str)
        .str.replace(r'\.0$', '', regex=True)
        .str.replace(r'\D', '', regex=True)
        .str.strip()
    )


def tratar_valor(coluna):
    return (
        coluna.astype(str)
        .str.replace("R$", "", regex=False)
        .str.replace(",", ".", regex=False)
        .str.strip()
    )


def detectar_coluna_codigo(df):
    for c in df.columns:
        if 'codigo' in c or 'ean' in c:
            return c
    return None


def detectar_coluna_preco(df):
    for c in df.columns:
        if 'preco' in c or 'valor' in c:
            return c
    return df.columns[-1]


# ==============================
# CARREGAR FORNECEDORES
# ==============================
def carregar_fornecedores(pasta):
    arquivos = [f for f in os.listdir(pasta) if f.endswith('.xlsx')]
    dfs = []

    for arquivo in arquivos:
        caminho = os.path.join(pasta, arquivo)
        df = pd.read_excel(caminho, dtype=str)

        df = tratar_colunas(df)

        col_codigo = detectar_coluna_codigo(df)

        if not col_codigo:
            print(f"⚠️ {arquivo} ignorado (sem código)")
            continue

        df = df.rename(columns={col_codigo: 'codigo'})
        df['codigo'] = tratar_codigo(df['codigo'])

        col_preco = detectar_coluna_preco(df)

        df['preco'] = tratar_valor(df[col_preco])
        df['preco'] = pd.to_numeric(df['preco'], errors='coerce')

        # REMOVE DUPLICAÇÃO
        df = df.groupby('codigo', as_index=False).agg({'preco': 'min'})

        nome = arquivo.replace('.xlsx', '').upper()
        df = df.rename(columns={'preco': nome})

        dfs.append(df)

    if not dfs:
        return pd.DataFrame()

    df_final = dfs[0]

    for df in dfs[1:]:
        df_final = df_final.merge(df, on='codigo', how='outer')

    return df_final


# ==============================
# CALCULAR MELHORES OPÇÕES
# ==============================
def calcular_melhores_opcoes(df, colunas_precos):

    def melhor(row):
        precos = row[colunas_precos].dropna().sort_values()
        return (precos.iloc[0], precos.index[0]) if len(precos) > 0 else (None, None)

    def segundo(row):
        precos = row[colunas_precos].dropna().sort_values()
        return (precos.iloc[1], precos.index[1]) if len(precos) > 1 else (None, None)

    df[['menor_preco', 'fornecedor_menor']] = df.apply(lambda r: pd.Series(melhor(r)), axis=1)
    df[['segundo_preco', 'fornecedor_segundo']] = df.apply(lambda r: pd.Series(segundo(r)), axis=1)

    return df


# ==============================
# PROCESSAMENTO
# ==============================
def processar_dados(df_base, df_fornecedores):

    df_base['codigo'] = tratar_codigo(df_base['codigo']).astype(str)
    df_fornecedores['codigo'] = tratar_codigo(df_fornecedores['codigo']).astype(str)

    df = df_base.merge(df_fornecedores, on='codigo', how='left')

    # remove duplicadas
    df = df.loc[:, ~df.columns.duplicated()]

    colunas_base = [
        'codigo', 'descricao.1', 'apresentacao',
        'laboratorio', 'quantidade', 'ultimo_valor_pago'
    ]

    colunas_preco = [c for c in df.columns if c not in colunas_base]

    for col in colunas_preco:
        df[col] = pd.to_numeric(df[col], errors='coerce')

    df = calcular_melhores_opcoes(df, colunas_preco)

    # 🔥 CORREÇÃO DO ERRO AQUI
    df['quantidade'] = (
        df['quantidade']
        .astype(str)
        .str.replace(",", ".")
    )

    df['quantidade'] = pd.to_numeric(df['quantidade'], errors='coerce')
    df['menor_preco'] = pd.to_numeric(df['menor_preco'], errors='coerce')
    df['segundo_preco'] = pd.to_numeric(df['segundo_preco'], errors='coerce')

    df['total_menor'] = df['menor_preco'] * df['quantidade']
    df['total_segundo'] = df['segundo_preco'] * df['quantidade']

    return df


# ==============================
# FORMATAÇÃO EXCEL
# ==============================
def colorir_precos_maiores(ws):
    fill_vermelho = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    colunas = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    col_ultimo = colunas.get('ultimo_valor_pago')

    if not col_ultimo:
        return

    colunas_base = [
        'codigo', 'descricao', 'quantidade', 'ultimo_valor_pago',
        'menor_preco', 'fornecedor_menor',
        'segundo_preco', 'fornecedor_segundo',
        'total_menor', 'total_segundo'
    ]

    colunas_fornecedores = [
        (nome, idx) for nome, idx in colunas.items()
        if nome not in colunas_base
    ]

    for row in range(2, ws.max_row + 1):
        ultimo_valor = ws.cell(row=row, column=col_ultimo).value

        if ultimo_valor is None:
            continue

        for nome, col_idx in colunas_fornecedores:
            valor = ws.cell(row=row, column=col_idx).value

            try:
                if valor is not None and float(valor) > float(ultimo_valor):
                    ws.cell(row=row, column=col_idx).fill = fill_vermelho
            except:
                pass


def formatar_excel(caminho):
    wb = load_workbook(caminho)
    ws = wb.active

    fill_header = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    font_header = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal='center')

    colorir_precos_maiores(ws)

    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 20

    wb.save(caminho)


def exportar_excel(df, caminho_saida):
    df.to_excel(caminho_saida, index=False)
    formatar_excel(caminho_saida)


# ==============================
# MAIN
# ==============================
def main():
    print("📊 Carregando base...")
    df_base = pd.read_excel("base_produtos.xlsx", dtype=str)
    df_base = tratar_colunas(df_base)

    col_codigo = detectar_coluna_codigo(df_base)

    if not col_codigo:
        raise Exception("❌ Base sem coluna de código")

    df_base = df_base.rename(columns={col_codigo: 'codigo'})
    df_base['codigo'] = tratar_codigo(df_base['codigo'])

    print("📦 Carregando fornecedores...")
    df_fornecedores = carregar_fornecedores("./fornecedores")

    print("⚙️ Processando...")
    df_final = processar_dados(df_base, df_fornecedores)

    print("📁 Exportando...")
    exportar_excel(df_final, "cotacao_final.xlsx")

    print("✅ Concluído com sucesso!")


if __name__ == "__main__":
    main()